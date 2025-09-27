import os
import time
import pandas as pd
from mpdptw.common.cli import parse_instance_argv
from mpdptw.common.printers.two_index_solution_printer import print_solution_summary
from mpdptw.common.parsers import build_milp_data
from gurobipy import *
import re 
from mpdptw.common.big_M import tight_bigM

def Run_Model(path, model: Model):
    start_time = time.perf_counter()
    inst = build_milp_data(str(path))




    EPS = 1e-6
    # Sets (extended with sink depot)
    V = inst["V_ext"]  # all nodes including origin (0) and sink
    A = inst["A_feasible_ext"]  # feasible arcs 
    N = inst["N"]

    R = inst["R"]  # request ids
    Pr = inst["Pr"]  # pickups per request
    Dr = inst["Dr"]  # deliveries per request
    Dr_single = inst["Dr_single"]  # single delivery per request
    S_min = inst["S_minimal_ext"]  # minimal S-sets from paper

    # Build adjacency from A once 
    in_arcs  = {j: [] for j in V}
    out_arcs = {i: [] for i in V}
    for (i, j) in A:
        out_arcs[i].append((i, j))
        in_arcs[j].append((i, j))
    
    K = inst["K"]  # vehicles
    Q = inst["Q"]  # capacity

    # Parameters (extended)
    e = inst["e"]  # earliest start
    l = inst["l"]  # latest start

    d = inst["d"]  # service time
    q = inst["q"]  # demand
    t = inst["t_ext"]  # travel time (extended)
    c = inst["c_ext"]  # travel cost (extended)
    V_ext = inst["V_ext"]

    # Special nodes
    depot = inst["depot"]  # start depot (0)
    sink = inst["sink"]  # sink depot
    #M_ij = {(i,j): max(0.0, l[i] + d[i] + t[i,j] - e[j]) for (i,j) in A}
    M_ij, Earliest, Latest = tight_bigM(out_arcs, t, d, V, A, sink, e, l, Pr=Pr, Dr_single=Dr_single)
    Pickups = [i for r in R for i in Pr[r]]
    Dels = [Dr_single[r] for r in R]
    N_only = set(N)
    # Map node -> request (for lifted SECs / precedence)
    node_req = {}
    for r in R:
        for p in Pr[r]:
            node_req[p] = r
        node_req[Dr_single[r]] = r
    node_type = {"delivery": set(Dels), "pickup": set(Pickups)}

    start_nodes = [j for (_, j) in out_arcs[depot] if j != sink]
    # Decision variables (to be declared in solver)
    X = {(i, j): model.addVar(vtype=GRB.BINARY) for (i, j) in A}  # binary arc use
    S = {i: model.addVar(vtype=GRB.CONTINUOUS, lb=Earliest[i], ub=Latest[i]) for i in V}  # continuous service start times



    model.setObjective(quicksum(X[i, j] * c[i, j] for (i, j) in A), GRB.MINIMIZE)


    # Degree (incoming = 1 for each customer j)
    DegreeConstrainIncome = {
        j: model.addConstr(quicksum(X[i, j] for (i, _) in in_arcs[j]) == 1)
        for j in N
    }

    # Degree (outgoing = 1 for each customer i)
    DegreeConstrainOutgoing = {
        i: model.addConstr(quicksum(X[i, j] for (_, j) in out_arcs[i]) == 1)
        for i in N
    }

    DepotBalance =  model.addConstr(
        quicksum(X[i, sink] for (i, _) in in_arcs[sink])
        == quicksum(X[0, j] for (_, j) in out_arcs[0])
    )

    TimeWindowFeas = {
        (i,j): model.addConstr(S[j] >= S[i] + d[i] + t[i,j] - M_ij[i,j] * (1 - X[i,j]))
        for (i,j) in A
    }
    TimeFeasEarliest = {i: model.addConstr(S[i] >= Earliest[i]) for i in V}
    TimeFeasLatest   = {i: model.addConstr(S[i] <= Latest[i])   for i in V}

    VehicleUsageLimit = model.addConstr(quicksum(X[0, j] for (_,j) in out_arcs[0]) <= len(K))

    RequestPrec = {(i, r):
                   model.addConstr(S[Dr_single[r]] >= S[i] + d[i] + t[i, Dr_single[r]])
                   for r in R for i in Pr[r]}
    


    def build_Sset(xvals):
        """Return internal cycles entirely within N, using hard 0/1 rounding of X."""
        Sset = set()
        succ = {}
        for i in N_only:
            outs = [(j, xvals.get((i, j), 0.0)) for (ii, j) in out_arcs.get(i, []) if j in N_only]
            if not outs:
                continue
            j_best, v_best = max(outs, key=lambda z: z[1])
            if v_best >= 1.0 - EPS:
                succ[i] = j_best

        unvisited = set(N_only)
        while unvisited:
            u = unvisited.pop()
            path, pos = [], {}
            while True:
                if u not in succ:
                    break
                if u in pos:
                    k = pos[u]
                    cycle = path[k:]
                    if len(cycle) >= 2:
                        Sset.add(frozenset(cycle))
                    break
                pos[u] = len(path)
                path.append(u)
                unvisited.discard(u)
                u = succ[u]
        return Sset

    def build_pi_sigma(S):
        """
        Compute π(S) and σ(S) for MPDPTW lifted SECs.

        Returns
        -------
        (pi_S, sigma_S) : tuple[set, set]
            pi_S    : union of pickup nodes for requests whose delivery ∈ S
            sigma_S : set of delivery nodes for requests having ≥1 pickup ∈ S
        """
        deliveries_in_S = S & node_type["delivery"]
        pickups_in_S = S & node_type["pickup"]

        reqs_with_delivery_in_S = {node_req[n] for n in deliveries_in_S if n in node_req}
        reqs_with_pickup_in_S = {node_req[n] for n in pickups_in_S if n in node_req}

        if reqs_with_delivery_in_S:
            pi_S = {n for r in reqs_with_delivery_in_S for n in Pr.get(r, ())}
        else:
            pi_S = set()

        sigma_S = {Dr_single[r] for r in reqs_with_pickup_in_S if r in Dr_single}
        return pi_S, sigma_S

    model._sec_seen = set()
    model._seen = set()
    model._have_incumbent = False  
    def subtour_callback(model, where):
        if where != GRB.Callback.MIPSOL:
            return
        XV = model.cbGetSolution(X)

        def sum_x_within(S):
            return quicksum(X[i, j] for i in S for j in S if (i, j) in X)

        def sum_x(from_set, to_set):
            return quicksum(X[i, j] for i in from_set for j in to_set if (i, j) in X)

        def add_cut_once(key, expr, rhs):
            if key in model._seen:
                return False
            model._seen.add(key)
            model.cbLazy(expr <= rhs)
            return True

        def sum_x_within_val(S, XV_):
            return sum(XV_.get((i, j), 0.0) for i in S for j in S)

        def sum_x_val(F, T, XV_):
            return sum(XV_.get((i, j), 0.0) for i in F for j in T)

        # ---- SECs on internal components ----
        cut_added = False
        Sset = build_Sset(XV)

        for S in Sset:
            S = set(S)
            Sbar = N_only - S
            pi_S, sigma_S = build_pi_sigma(S)  # π(S), σ(S)
            rhs = len(S) - 1

            # σ-inequality (49)
            lhs_sigma_expr = sum_x_within(S) + sum_x(Sbar & sigma_S, S) + sum_x(Sbar - sigma_S, S & sigma_S)
            lhs_sigma_val = (
                sum_x_within_val(S, XV)
                + sum_x_val(Sbar & sigma_S, S, XV)
                + sum_x_val(Sbar - sigma_S, S & sigma_S, XV)
            )
            if lhs_sigma_val > rhs + EPS:
                key = ("LSEC_sigma", frozenset(S))
                cut_added |= add_cut_once(key, lhs_sigma_expr, rhs)

            # π-inequality (50)
            lhs_pi_expr = sum_x_within(S) + sum_x(S, Sbar & pi_S) + sum_x(S & pi_S, Sbar - pi_S)
            lhs_pi_val = (
                sum_x_within_val(S, XV)
                + sum_x_val(S, Sbar & pi_S, XV)
                + sum_x_val(S & pi_S, Sbar - pi_S, XV)
            )
            if lhs_pi_val > rhs + EPS:
                key = ("LSEC_pi", frozenset(S))
                cut_added |= add_cut_once(key, lhs_pi_expr, rhs)

        # ---- Route-based precedence cuts (apply to ALL routes from the depot) ----
        start_arcs = [arc for arc in out_arcs[depot] if XV.get(arc, 0.0) > 0.5]

        for i0, j0 in start_arcs:
            # Follow successors until sink or repetition
            route = [i0, j0]
            posr = {i0: 0, j0: 1}
            seen = {i0, j0}
            node = j0
            for _ in range(len(V) + 2):
                if node == sink:
                    break
                succs = [j for (_, j) in out_arcs.get(node, []) if XV.get((node, j), 0.0) > 0.5]
                if not succs:
                    break
                j = succs[0]
                if j in seen:
                    route.append(j)
                    posr[j] = len(route) - 1
                    break
                route.append(j)
                posr[j] = len(route) - 1
                seen.add(j)
                node = j

            reqs_on_route = {node_req[v] for v in route if v in node_req}

            # (54): delivery before all pickups of same request
            for r in reqs_on_route:
                dr = Dr_single[r]
                if dr not in posr:
                    continue
                idx_d = posr[dr]
                for p in Pr[r]:
                    if p in posr and posr[p] > idx_d:
                        S_list = route[idx_d + 1 : posr[p]]
                        if not S_list:
                            continue
                        S = set(S_list)
                        key = ("54", dr, p, tuple(S_list))
                        expr = sum_x({dr}, S) + sum_x_within(S) + sum_x(S, {p})
                        cut_added |= add_cut_once(key, expr, len(S))

            # (52): depot → ... → dr but some pickup is not before dr (incl. pickup on another route)
            for r in reqs_on_route:
                dr = Dr_single[r]
                if dr not in posr:
                    continue
                idx_d = posr[dr]
                late_or_missing = any((p not in posr) or (posr[p] > idx_d) for p in Pr[r])
                if not late_or_missing:
                    continue
                S_list = route[1:idx_d]  # strictly between depot and dr
                if not S_list:
                    continue
                S = set(S_list)
                key = ("52", dr, tuple(S_list))
                expr = sum_x({depot}, S) + sum_x_within(S | {dr})
                cut_added |= add_cut_once(key, expr, len(S))

            # (53): pickup -> ... -> depot, but dr is not between p and depot (incl. dr on another route)
            for r in reqs_on_route:
                dr = Dr_single[r]
                for p in Pr[r]:
                    if p not in posr:
                        continue
                    idx_p = posr[p]
                    idx_sink = posr.get(sink)
                    if idx_sink is None or idx_sink <= idx_p:
                        continue
                    dr_after_p = (dr in posr) and (idx_p < posr[dr] < idx_sink)
                    if dr_after_p:
                        continue
                    S_list = route[idx_p + 1 : idx_sink]  # strictly between p and depot
                    if not S_list:
                        continue
                    S = set(S_list)
                    key = ("53", p, tuple(S_list))
                    expr = sum_x_within(S | {p}) + sum_x(S, {sink})
                    cut_added |= add_cut_once(key, expr, len(S))
    model.setParam('TimeLimit', 1800)
    model.optimize(subtour_callback)
    end_time = time.perf_counter()
    # load UB table once globally
    ub_df = pd.read_csv(r"mpdtw_instances_2019\upper_bounds.csv")  # Instance, UB
    UB_TABLE = dict(zip(ub_df.Instance, ub_df.UB))

    # --- after optimize() ---
    has_incumbent = (model.SolCount > 0)
    is_optimal = (model.Status == GRB.OPTIMAL)

    LB = model.ObjBound
    elapsed_time = end_time - start_time
    if elapsed_time > 1800:
        elapsed_time = 1800
    work_units = getattr(model, "Work", 0.0)

    filename = os.path.basename(path)
    name_without_ext = os.path.splitext(filename)[0]

    UB_file = UB_TABLE.get(name_without_ext)
    UB_inc = model.ObjVal if has_incumbent else None

    if has_incumbent and UB_file is not None:
        UB = min(UB_inc, UB_file)
    elif has_incumbent:
        UB = UB_inc
    else:
        UB = UB_file  # may be None

    MIP_Gap = None
    if UB is not None and UB > 0 and LB is not None:
        MIP_Gap = (UB - LB) / UB

    Sts = 1 if is_optimal else 0
    # --- write to Excel ---
    results_xlsx=r"docs\model_results.xlsx"
    try:
        update_two_index_block(
            xlsx_path=results_xlsx,
            instance_name=name_without_ext,
            Sts=Sts,
            Work=work_units,
            LB=LB,
            UB=UB,
            MIP_Gap=MIP_Gap,
            elapsed_time=elapsed_time,
        )
    except Exception as e:
        print(f"⚠ Could not write results for '{name_without_ext}': {e}")

def main(argv=None):
    folder = "mpdtw_instances_2019"

    filenames = [
        f for f in os.listdir(folder)
        if f.endswith(".txt")
        and f.startswith("l")
        and (len(f.split("_")) < 3 or f.split("_")[2].split(".")[0] in ("100"))
    ]

    filenames.sort(key=str.lower)
    print(filenames)
    total = len(filenames)

    for i, filename in enumerate(filenames, start=1):
        path = os.path.join(folder, filename)
        print(f"▶ Running ({i}/{total}) {filename}")

        model = Model("MPDTW")
        model.setParam(GRB.Param.LazyConstraints, 1)

        Run_Model(str(path), model)  # all metrics + Excel write happen inside

        print(f"✔ Done ({i}/{total}) {filename}\n")
        time.sleep(240)

if __name__ == "__main__":
    main()


from openpyxl import load_workbook
import re
from math import isfinite

def _norm(s):
    """Normalize cell text for matching: lowercase, collapse spaces, strip non-alnum except %."""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    # Keep % because it's meaningful here
    s = re.sub(r"[^a-z0-9 %\-\_()]", "", s)
    return s

def _find_cell(ws, target_text):
    """Find first cell whose normalized text equals target_text (normalized)."""
    tgt = _norm(target_text)
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if _norm(cell.value) == tgt:
                return cell.row, cell.column
    return None, None

def _find_instance_row(ws, instance_name):
    """Find the row index where a cell equals instance_name (normalized string match)."""
    tgt = _norm(instance_name)
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for v in row:
            if _norm(v) == tgt:
                return r
    return None

def _find_subheader_columns(ws, header_row, start_col, expected_map):
    """
    expected_map: dict of canonical keys -> list of acceptable header variants.
    Returns dict key -> column index.
    """
    found = {}
    # Scan to the right for a reasonable span (e.g., 30 cols)
    for c in range(start_col, start_col + 30):
        label = _norm(ws.cell(row=header_row, column=c).value)
        if not label:
            continue
        for key, variants in expected_map.items():
            if key in found:
                continue
            if label in variants:
                found[key] = c
    return found

def update_two_index_block(
    xlsx_path,
    instance_name,           # e.g., name_without_ext
    Sts, Work, LB, UB, MIP_Gap, elapsed_time,
):
    """
    Writes to the block:
        2-index model
            Sts | Work Units | LB | UB | Gap final % | Time (s)
    in the row matching 'instance_name'.
    - Sts: 0/1
    - Work: model.Work
    - LB, UB: bounds (UB can be inf -> cell left blank)
    - MIP_Gap: decimal (e.g., 0.1234). We'll store % (e.g., 12.34).
    - elapsed_time: seconds
    """
    wb = load_workbook(xlsx_path)
    # Try each sheet until we find the block
    for ws in wb.worksheets:
        r, c = _find_cell(ws, "2-index model")
        if r is None:
            continue

        header_row = r + 1  # subheaders are expected one row below the merged header
        # Acceptable normalized labels for each subcolumn
        expected = {
            "sts":            {"sts"},
            "work_units":     {"work units", "workunits", "work"},
            "lb":             {"lb"},
            "ub":             {"ub"},
            "gap_final_pct":  {"gap final %", "gap final%", "gap %", "gap", "mip gap %", "mip gap"},
            "time_s":         {"time (s)", "time s", "time", "time sec", "time seconds"},
        }
        # Normalize all variants once
        expected_map = {k: { _norm(v) for v in vals } for k, vals in expected.items()}

        cols = _find_subheader_columns(ws, header_row, c, expected_map)
        must_have = {"sts","work_units","lb","ub","gap_final_pct","time_s"}
        if not must_have.issubset(cols.keys()):
            # Couldn't map all subheaders on this sheet; try next sheet
            continue

        # Find the target row for the instance name anywhere in the sheet
        target_row = _find_instance_row(ws, instance_name)
        if target_row is None:
            # If not found, append a new row at the bottom, put the name in col A
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=1, value=instance_name)

        # Prepare values
        gap_pct = round(MIP_Gap * 100, 2) if MIP_Gap is not None else None
        val_map = {
            "sts":            int(Sts) if Sts is not None else None,
            "work_units":     round(Work, 2) if Work is not None else None,
            "lb":             round(LB, 2) if LB is not None else None,
            "ub":             (round(UB, 2) if (UB is not None and isfinite(UB)) else None),
            "gap_final_pct":  gap_pct,
            "time_s":         round(elapsed_time, 2) if elapsed_time is not None else None,
        }

        # Write values
        for key, col_idx in cols.items():
            ws.cell(row=target_row, column=col_idx, value=val_map[key])

        print(f"✔ Wrote results for '{instance_name}' into '2-index model' (row {target_row}): {val_map}")

        wb.save(xlsx_path)
        return True  # success

    # If we get here, we didn't find the block
    raise RuntimeError("Could not locate the '2-index model' block and its subheaders in the workbook.")


